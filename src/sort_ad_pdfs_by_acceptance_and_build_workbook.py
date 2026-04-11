"""
sort_ad_pdfs_by_acceptance_and_build_workbook.py
-------------------------------------------------
Step 2 of the AD Reproducibility Audit pipeline.

Parses PDFs of Alzheimer's & Dementia articles:
- Extracts acceptance date, DOI, and title from each PDF
- Sorts PDFs into month subfolders (January–December)
- Builds an OSF-style Excel workbook (one sheet per month)

Adapted from: KahinaBch/mrm-reproducible-research-2025
Original methodology: Boudreau et al. "On the open-source landscape of MRM"

Key differences from MRM pipeline:
- No author gender inference (dropped per project brief)
- Workbook columns adjusted accordingly
"""

import argparse
import re
import shutil
import logging
from pathlib import Path
from datetime import datetime

import pdfminer.high_level
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# OSF-style workbook columns — adapted for A&D (no gender columns)
WORKBOOK_COLUMNS = [
    "Filename",
    "DOI",
    "Title",
    "Acceptance Date",
    "Month",
    "Keywords Matched",
    "Data Availability Statement",
    "False Positive?",
    "Link",
    "Shared code?",
    "Shared data?",
    "Language(s)",
    "First author affiliation country",
    "Sex-specific keywords?",
    "Sex keywords matched",
    "Additional notes",
]

# ── Patterns ────────────────────────────────────────────────────────────────

ACCEPT_PATTERNS = [
    r"[Aa]ccepted[:\s]+(\d{1,2}\s+\w+\s+\d{4})",       # Accepted: 12 March 2023
    r"[Aa]ccepted[:\s]+(\w+\s+\d{1,2},?\s+\d{4})",      # Accepted: March 12, 2023
    r"[Aa]ccepted[:\s]+(\d{4}-\d{2}-\d{2})",             # Accepted: 2023-03-12
    r"[Rr]eceived.*?[Aa]ccepted[:\s]+(\d{1,2}\s+\w+\s+\d{4})",
]

DOI_PATTERN = re.compile(r"\b10\.\d{4,}/[^\s\"'<>]+", re.IGNORECASE)

MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def extract_text_first_pages(pdf_path: Path, max_pages: int = 4) -> str:
    """Extract text from the first N pages of a PDF."""
    try:
        text = pdfminer.high_level.extract_text(str(pdf_path), page_numbers=list(range(max_pages)))
        return text or ""
    except Exception as e:
        log.warning(f"  PDF parse error ({pdf_path.name}): {e}")
        return ""


def parse_acceptance_date(text: str) -> tuple[str, int | None]:
    """
    Try to extract acceptance date from PDF text.
    Returns (date_string, month_number) or ("", None).
    """
    for pattern in ACCEPT_PATTERNS:
        m = re.search(pattern, text)
        if m:
            raw = m.group(1).strip()
            # Try to determine month number
            for abbr, num in MONTH_MAP.items():
                if abbr in raw.lower():
                    return raw, num
            # Try YYYY-MM-DD
            iso = re.match(r"(\d{4})-(\d{2})-\d{2}", raw)
            if iso:
                return raw, int(iso.group(2))
    return "", None


def extract_doi(text: str) -> str:
    """Extract the first DOI found in text."""
    m = DOI_PATTERN.search(text)
    if m:
        doi = m.group(0).rstrip(".,;)")
        return doi
    return ""


def extract_title(text: str) -> str:
    """Heuristic: title is often in the first 500 chars, before abstract."""
    snippet = text[:500]
    lines = [l.strip() for l in snippet.splitlines() if len(l.strip()) > 20]
    return lines[0] if lines else ""


def sort_pdfs(pdf_folder: Path, year: int) -> dict[int, list[dict]]:
    """
    Parse all PDFs in folder, extract metadata, sort into month buckets.
    Returns dict: month_num -> list of record dicts.
    """
    month_buckets: dict[int, list[dict]] = {m: [] for m in range(1, 13)}
    unclassified = []

    pdfs = sorted(pdf_folder.glob("*.pdf"))
    log.info(f"Found {len(pdfs)} PDFs in {pdf_folder}")

    for pdf in pdfs:
        text = extract_text_first_pages(pdf, max_pages=3)
        date_str, month_num = parse_acceptance_date(text)
        doi = extract_doi(text)
        title = extract_title(text)

        record = {
            "Filename": pdf.name,
            "DOI": doi,
            "Title": title,
            "Acceptance Date": date_str,
            "Month": MONTHS[month_num - 1] if month_num else "Unknown",
            "Keywords Matched": "",
            "Data Availability Statement": "",
            "False Positive?": "",
            "Link": "",
            "Shared code?": "",
            "Shared data?": "",
            "Language(s)": "",
            "First author affiliation country": "",
            "Sex-specific keywords?": "",
            "Sex keywords matched": "",
            "Additional notes": "",
            "_pdf_path": pdf,
        }

        if month_num:
            month_buckets[month_num].append(record)
            # Move PDF to month subfolder
            month_folder = pdf_folder / MONTHS[month_num - 1]
            month_folder.mkdir(exist_ok=True)
            dest = month_folder / pdf.name
            if not dest.exists():
                shutil.copy2(pdf, dest)
        else:
            log.warning(f"  Could not determine acceptance month for: {pdf.name}")
            unclassified.append(record)

    log.info(f"  Classified: {sum(len(v) for v in month_buckets.values())} PDFs")
    log.info(f"  Unclassified: {len(unclassified)} PDFs")

    # Put unclassified in month 13 for inspection
    month_buckets[13] = unclassified
    return month_buckets


def build_workbook(month_buckets: dict, year: int, out_path: Path):
    """Build an OSF-style Excel workbook with one sheet per month."""
    wb = openpyxl.Workbook()

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    header_fill = PatternFill("solid", fgColor="6B21A8")  # Purple
    alt_fill = PatternFill("solid", fgColor="F3E8FF")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # Remove default sheet
    wb.remove(wb.active)

    for month_num in range(1, 13):
        records = month_buckets.get(month_num, [])
        month_name = MONTHS[month_num - 1]
        ws = wb.create_sheet(title=month_name)

        # Header
        for col_idx, col_name in enumerate(WORKBOOK_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_align

        # Data rows
        for row_idx, record in enumerate(records, start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, col_name in enumerate(WORKBOOK_COLUMNS, start=1):
                val = record.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = wrap_align
                if fill:
                    cell.fill = fill

        # Column widths
        col_widths = {
            "Filename": 30, "DOI": 25, "Title": 50, "Acceptance Date": 16,
            "Month": 12, "Keywords Matched": 30, "Data Availability Statement": 40,
            "False Positive?": 14, "Link": 30, "Shared code?": 13,
            "Shared data?": 13, "Language(s)": 15,
            "First author affiliation country": 20,
            "Sex-specific keywords?": 18, "Sex keywords matched": 30,
            "Additional notes": 40,
        }
        for col_idx, col_name in enumerate(WORKBOOK_COLUMNS, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = col_widths.get(col_name, 15)

        ws.freeze_panes = "A2"

        log.info(f"  Sheet '{month_name}': {len(records)} papers")

    # Unclassified sheet
    unclassified = month_buckets.get(13, [])
    if unclassified:
        ws_unc = wb.create_sheet(title="Unclassified")
        for col_idx, col_name in enumerate(WORKBOOK_COLUMNS, start=1):
            cell = ws_unc.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor="BE185D")
            cell.alignment = wrap_align
        for row_idx, record in enumerate(unclassified, start=2):
            for col_idx, col_name in enumerate(WORKBOOK_COLUMNS, start=1):
                ws_unc.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info(f"\nWorkbook saved: {out_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Sort A&D PDFs by acceptance month and build OSF-style workbook."
    )
    parser.add_argument("--year", type=int, required=True, help="Year to process")
    parser.add_argument("--pdf-folder", type=Path, required=True, help="Folder containing downloaded PDFs")
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Output workbook path (default: workbooks/{year}/AD-ReproducibleResearch_{year}.xlsx)",
    )
    args = parser.parse_args()

    out_path = args.out or Path(f"workbooks/{args.year}/AD-ReproducibleResearch_{args.year}.xlsx")

    log.info(f"=== Step 2: Sort PDFs and build workbook | Year: {args.year} ===")
    month_buckets = sort_pdfs(args.pdf_folder, args.year)
    build_workbook(month_buckets, args.year, out_path)
    log.info("=== Step 2 complete ===")


if __name__ == "__main__":
    main()
