"""
add_author_gender_from_doi.py
--------------------------------
Intermediate step for the AD Reproducibility Audit pipeline.

For each paper (via DOI), this script:
- Retrieves author metadata from Crossref
- Adds first and last author names to workbook
- Infers first-name gender when identifiable

Workbook columns added/updated:
- First author
- First author gender
- Last author
- Last author gender
"""

from __future__ import annotations

import argparse
import csv
import logging
import time
from pathlib import Path
from urllib.parse import quote

import openpyxl
import requests
from gender_guesser import detector as gender_detector

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


def map_gender(label: str) -> str:
    label = (label or "").lower()
    if label in {"male", "mostly_male"}:
        return "Male"
    if label in {"female", "mostly_female"}:
        return "Female"
    if label == "andy":
        return "Androgynous"
    return "Unknown"


def infer_gender_from_name(full_name: str, detector: gender_detector.Detector) -> str:
    if not full_name:
        return "Unknown"
    first_token = full_name.split()[0].replace(".", "").replace("-", " ").split()[0].strip()
    if len(first_token) < 2:
        return "Unknown"
    return map_gender(detector.get_gender(first_token))


def author_display_name(author_obj: dict) -> str:
    given = (author_obj.get("given") or "").strip()
    family = (author_obj.get("family") or "").strip()
    name = " ".join([x for x in [given, family] if x]).strip()
    if name:
        return name
    return (author_obj.get("name") or "").strip()


def fetch_authors_from_crossref(session: requests.Session, doi: str) -> list[dict]:
    doi_clean = (doi or "").strip()
    if not doi_clean:
        return []

    url = f"https://api.crossref.org/works/{quote(doi_clean, safe='') }"

    for attempt in range(3):
        try:
            r = session.get(url, timeout=20)
            if r.status_code == 200:
                msg = (r.json() or {}).get("message") or {}
                return msg.get("author") or []
            if r.status_code in {404, 422}:
                return []
            if r.status_code == 429:
                time.sleep(1.2 * (attempt + 1))
                continue
            time.sleep(0.4 * (attempt + 1))
        except Exception:
            time.sleep(0.4 * (attempt + 1))
    return []


def main() -> None:
    parser = argparse.ArgumentParser(description="Add first/last author names and inferred genders from DOI metadata.")
    parser.add_argument("--xlsx", type=Path, required=True, help="Path to workbook")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx)
    session = requests.Session()
    session.headers.update({"User-Agent": "ad-reproducibility-audit/1.0 (mailto:kahina@example.com)"})
    detector = gender_detector.Detector(case_sensitive=False)

    cache: dict[str, tuple[str, str, str, str]] = {}
    log_rows: list[dict] = []

    total_rows = 0
    updated_rows = 0

    for month in MONTHS:
        if month not in wb.sheetnames:
            continue
        ws = wb[month]

        header = {cell.value: cell.column for cell in ws[1]}
        doi_col = header.get("DOI")
        if not doi_col:
            log.warning(f"  DOI column missing in '{month}' — skipping")
            continue

        # ensure target columns exist
        first_author_col = header.get("First author") or (ws.max_column + 1)
        if header.get("First author") is None:
            ws.cell(row=1, column=first_author_col, value="First author")

        first_gender_col = header.get("First author gender") or (ws.max_column + 1)
        if header.get("First author gender") is None:
            ws.cell(row=1, column=first_gender_col, value="First author gender")

        last_author_col = header.get("Last author") or (ws.max_column + 1)
        if header.get("Last author") is None:
            ws.cell(row=1, column=last_author_col, value="Last author")

        last_gender_col = header.get("Last author gender") or (ws.max_column + 1)
        if header.get("Last author gender") is None:
            ws.cell(row=1, column=last_gender_col, value="Last author gender")

        for row_idx in range(2, ws.max_row + 1):
            doi = (ws.cell(row=row_idx, column=doi_col).value or "").strip() if ws.cell(row=row_idx, column=doi_col).value else ""
            total_rows += 1

            if not doi:
                log_rows.append({
                    "month": month,
                    "row": row_idx,
                    "doi": "",
                    "first_author": "",
                    "first_author_gender": "Unknown",
                    "last_author": "",
                    "last_author_gender": "Unknown",
                    "status": "missing_doi",
                })
                continue

            if doi in cache:
                fa, fag, la, lag = cache[doi]
            else:
                authors = fetch_authors_from_crossref(session, doi)
                if not authors:
                    fa = la = ""
                    fag = lag = "Unknown"
                else:
                    first = authors[0]
                    last = authors[-1]
                    fa = author_display_name(first)
                    la = author_display_name(last)
                    fag = infer_gender_from_name(fa, detector)
                    lag = infer_gender_from_name(la, detector)
                cache[doi] = (fa, fag, la, lag)

            ws.cell(row=row_idx, column=first_author_col, value=fa)
            ws.cell(row=row_idx, column=first_gender_col, value=fag)
            ws.cell(row=row_idx, column=last_author_col, value=la)
            ws.cell(row=row_idx, column=last_gender_col, value=lag)
            updated_rows += 1

            log_rows.append({
                "month": month,
                "row": row_idx,
                "doi": doi,
                "first_author": fa,
                "first_author_gender": fag,
                "last_author": la,
                "last_author_gender": lag,
                "status": "ok" if fa or la else "no_author_metadata",
            })

            time.sleep(0.05)

    wb.save(args.xlsx)

    log_path = args.xlsx.parent / "author_gender_log.csv"
    with open(log_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "month",
                "row",
                "doi",
                "first_author",
                "first_author_gender",
                "last_author",
                "last_author_gender",
                "status",
            ],
        )
        writer.writeheader()
        writer.writerows(log_rows)

    log.info("=== Intermediate step complete: author + gender enrichment ===")
    log.info(f"  Rows visited : {total_rows}")
    log.info(f"  Rows updated : {updated_rows}")
    log.info(f"  DOI cache size: {len(cache)}")
    log.info(f"  Log saved    : {log_path}")


if __name__ == "__main__":
    main()
