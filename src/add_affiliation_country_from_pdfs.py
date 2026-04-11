"""
add_affiliation_country_from_pdfs.py
--------------------------------------
Step 5 of the AD Reproducibility Audit pipeline.

Extracts the first author's affiliation country directly from PDF text,
then updates the workbook column "First author affiliation country".

Strategy (adapted from MRM pipeline):
1. Build DOI → PDF path index by scanning first pages of each PDF
2. Extract first-page text using pdfminer (better layout than PyPDF2)
3. Restrict to pre-Abstract region (affiliations are in header)
4. Heuristically identify the first affiliation line
5. Infer country using pycountry + curated alias dictionary

Adapted from: KahinaBch/mrm-reproducible-research-2025
"""

import argparse
import csv
import logging
import re
from pathlib import Path

import pdfminer.high_level
import openpyxl
import pycountry

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# ── Country alias dictionary ─────────────────────────────────────────────────
# Maps common aliases / abbreviations found in affiliations to pycountry names
COUNTRY_ALIASES = {
    "usa": "United States",
    "u.s.a.": "United States",
    "u.s.": "United States",
    "united states": "United States",
    "uk": "United Kingdom",
    "u.k.": "United Kingdom",
    "england": "United Kingdom",
    "scotland": "United Kingdom",
    "wales": "United Kingdom",
    "great britain": "United Kingdom",
    "p.r. china": "China",
    "pr china": "China",
    "people's republic of china": "China",
    "roc": "Taiwan",
    "south korea": "Korea, Republic of",
    "republic of korea": "Korea, Republic of",
    "korea": "Korea, Republic of",
    "the netherlands": "Netherlands",
    "holland": "Netherlands",
    "uae": "United Arab Emirates",
    "brasil": "Brazil",
    "deutschland": "Germany",
    "espana": "Spain",
    "españa": "Spain",
    "italia": "Italy",
}

# Stopwords that appear in affiliations but are NOT countries
NOT_COUNTRIES = {
    "medicine", "neurology", "sciences", "university", "hospital",
    "institute", "center", "centre", "department", "school",
    "college", "faculty", "division", "laboratory", "clinic",
}


def extract_first_page_text(pdf_path: Path) -> str:
    """Extract text from the first page only using pdfminer."""
    try:
        return pdfminer.high_level.extract_text(str(pdf_path), page_numbers=[0]) or ""
    except Exception as e:
        log.warning(f"  pdfminer failed on {pdf_path.name}: {e}")
        return ""


def get_pre_abstract_region(text: str) -> str:
    """
    Return only the text before 'Abstract' / 'Introduction' heading.
    Affiliations appear in this region.
    """
    for marker in ["Abstract\n", "ABSTRACT\n", "Introduction\n", "INTRODUCTION\n",
                   "Background\n", "BACKGROUND\n"]:
        idx = text.find(marker)
        if idx > 50:
            return text[:idx]
    return text[:1500]  # Fallback: first 1500 chars


def infer_country(text: str) -> str:
    """
    Infer country from an affiliation text fragment.

    Strategy:
    1. Check alias dictionary (case-insensitive)
    2. Try pycountry lookup against each comma/semicolon-separated token
    3. Return first match found

    Returns country name string or "" if not found.
    """
    # Normalise
    text_lower = text.lower()

    # 1. Alias check
    for alias, country in COUNTRY_ALIASES.items():
        if alias in text_lower:
            return country

    # 2. Token-by-token pycountry lookup
    tokens = re.split(r"[,;\n]+", text)
    for token in reversed(tokens):  # Country usually last in affiliation string
        token = token.strip().strip(".")
        if not token or len(token) < 3:
            continue
        if token.lower() in NOT_COUNTRIES:
            continue
        # Try exact match
        match = pycountry.countries.get(name=token)
        if match:
            return match.name
        # Try fuzzy search (limited to avoid false positives)
        try:
            results = pycountry.countries.search_fuzzy(token)
            if results and len(token) > 5:
                return results[0].name
        except LookupError:
            pass

    return ""


def build_doi_to_pdf_index(year_folder: Path) -> dict[str, Path]:
    """
    Build a mapping DOI → PDF path by scanning first pages of all PDFs.
    """
    doi_pattern = re.compile(r"\b10\.\d{4,}/[^\s\"'<>]+", re.IGNORECASE)
    index = {}

    for pdf in year_folder.rglob("*.pdf"):
        text = extract_first_page_text(pdf)
        m = doi_pattern.search(text)
        if m:
            doi = m.group(0).rstrip(".,;)").lower()
            index[doi] = pdf

    log.info(f"  DOI index built: {len(index)} PDFs matched to DOIs")
    return index


def update_workbook_countries(
    year_folder: Path,
    workbook_path: Path,
    log_rows: list[dict],
):
    """Extract affiliation country for each paper and update workbook."""
    doi_index = build_doi_to_pdf_index(year_folder)

    # Also build filename → path index as fallback
    fn_index = {pdf.name.lower(): pdf for pdf in year_folder.rglob("*.pdf")}

    wb = openpyxl.load_workbook(workbook_path)
    # Create backup
    backup_path = workbook_path.with_suffix(".backup.xlsx")
    wb.save(backup_path)

    for month in MONTHS:
        if month not in wb.sheetnames:
            continue

        ws = wb[month]
        header = {cell.value: cell.column for cell in ws[1]}

        country_col = header.get("First author affiliation country")
        doi_col = header.get("DOI")
        fn_col = header.get("Filename")

        if not country_col:
            log.warning(f"  Country column missing in sheet '{month}' — skipping.")
            continue

        for row_idx in range(2, ws.max_row + 1):
            doi_val = (ws.cell(row=row_idx, column=doi_col).value or "").lower().strip() if doi_col else ""
            fn_val = (ws.cell(row=row_idx, column=fn_col).value or "").lower().strip() if fn_col else ""

            # Find PDF
            pdf_path = doi_index.get(doi_val) or fn_index.get(fn_val)
            if not pdf_path:
                log.debug(f"    No PDF found for row {row_idx} in {month}")
                log_rows.append({"month": month, "row": row_idx, "doi": doi_val, "pdf": fn_val,
                                  "country": "", "status": "pdf_not_found"})
                continue

            text = extract_first_page_text(pdf_path)
            region = get_pre_abstract_region(text)
            country = infer_country(region)

            ws.cell(row=row_idx, column=country_col, value=country)
            log_rows.append({
                "month": month, "row": row_idx, "doi": doi_val,
                "pdf": pdf_path.name, "country": country,
                "status": "ok" if country else "country_not_found",
            })

            if not country:
                log.debug(f"    Country not found: {pdf_path.name}")

    wb.save(workbook_path)
    log.info(f"\nWorkbook updated with country data.")


def main():
    parser = argparse.ArgumentParser(
        description="Extract first-author affiliation country from PDFs."
    )
    parser.add_argument("--year-folder", type=Path, required=True,
                        help="Folder containing month subfolders with PDFs")
    parser.add_argument("--xlsx", type=Path, required=True,
                        help="Path to the Excel workbook")
    args = parser.parse_args()

    log.info("=== Step 5: Country extraction from PDFs ===")
    log_rows: list[dict] = []

    update_workbook_countries(args.year_folder, args.xlsx, log_rows)

    log_path = args.xlsx.parent / "pdf_affiliation_country_log.csv"
    with open(log_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["month", "row", "doi", "pdf", "country", "status"])
        writer.writeheader()
        writer.writerows(log_rows)

    total = len(log_rows)
    found = sum(1 for r in log_rows if r["country"])
    log.info(f"\nCountry extraction summary:")
    log.info(f"  Rows processed : {total}")
    log.info(f"  Countries found: {found} ({100*found/total:.1f}%)" if total else "  No rows processed.")
    log.info(f"  Log saved      : {log_path}")
    log.info("=== Step 5 complete ===")


if __name__ == "__main__":
    main()
