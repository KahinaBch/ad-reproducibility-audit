"""
scan_sex_keywords_update_workbook.py
--------------------------------------
Step 4 of the AD Reproducibility Audit pipeline.

NOVEL CONTRIBUTION — Not present in the MRM pipeline.

Scans PDFs for keywords indicating sex-specific or sex-stratified analyses.
Updates the workbook columns:
  - "Sex-specific keywords?" (Yes / No)
  - "Sex keywords matched"   (list of matched keywords)
    - "Sex-aware level"       (sex-aware main focus / sex-aware consideration)

Scientific rationale:
  Despite women constituting ~65% of people living with Alzheimer's disease,
  sex-stratified analyses remain rare in the AD literature. This step quantifies
  the proportion of papers that explicitly account for sex as a biological variable.

Keywords validated by the project team (GBM6330E, 2025):
  Core (confirmed): sex-stratified, sex differences, gender-specific
  Extended (approved): sex-disaggregated, sex-based analysis, female-specific,
                       sex as a biological variable, sex as a covariate,
                       hormonal, menopause, APOE sex interaction,
                       sex-conditioned, male-specific

Adapted from: KahinaBch/mrm-reproducible-research-2025 (keyword scan logic)
"""

import argparse
import csv
import logging
from pathlib import Path

import openpyxl
import PyPDF2

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Sex-specific keyword set ─────────────────────────────────────────────────
SEX_KEYWORDS = [
    # Broad terms requested by user
    "sex",
    "gender",
    "woman",
    "female",
    # Core — confirmed by user
    "sex-stratified",
    "sex stratified",
    "sex differences",
    "gender-specific",
    "gender specific",
    # Extended — approved by user
    "sex-disaggregated",
    "sex disaggregated",
    "sex-based analysis",
    "sex based analysis",
    "female-specific",
    "female specific",
    "male-specific",
    "male specific",
    "sex as a biological variable",
    "sex as a covariate",
    "sex-conditioned",
    "sex conditioned",
    # Biological context keywords (AD-specific)
    "menopause",
    "hormonal influence",
    "hormonal factors",
    "apoe sex interaction",
    "sex interaction",
    "estrogen",
    "testosterone",
    "sex-stratified analysis",
    "stratified by sex",
    "stratified by gender",
    "subgroup analysis by sex",
    "subgroup analysis by gender",
    "sex-specific",
    "sex specific",
]

# These broad terms are intentionally checked only in titles
TITLE_ONLY_KEYWORDS = {
    "sex",
    "gender",
    "woman",
    "female",
}

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


def extract_full_text_pypdf2(pdf_path: Path) -> str:
    """Extract full document text using PyPDF2."""
    full_text = []
    try:
        with open(pdf_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                try:
                    full_text.append(page.extract_text() or "")
                except Exception:
                    full_text.append("")
    except Exception as e:
        log.warning(f"  PyPDF2 failed on {pdf_path.name}: {e}")
    return "\n".join(full_text)


def detect_sex_keywords(text: str, keywords: list[str] | None = None) -> list[str]:
    """
    Detect sex-specific keywords in document text.
    Returns sorted list of matched keywords (deduplicated).
    Note: scans full document (unlike open-science scan) to capture Methods,
    Results, and Discussion sections where sex analysis would be reported.
    """
    text_lower = text.lower()
    matched = set()
    key_list = keywords if keywords is not None else SEX_KEYWORDS
    for kw in key_list:
        if kw.lower() in text_lower:
            matched.add(kw)
    return sorted(matched)


def classify_sex_aware_level(title_text: str, full_text: str) -> tuple[str, list[str], list[str]]:
    """
    Classify paper level:
      - sex-aware main focus: keyword appears in title
      - sex-aware consideration: keyword appears in main text (but not title)

    Returns: (level, title_matched, full_matched)
    """
    title_matched = detect_sex_keywords(title_text or "")
    fulltext_keywords = [kw for kw in SEX_KEYWORDS if kw not in TITLE_ONLY_KEYWORDS]
    full_matched = detect_sex_keywords(full_text or "", keywords=fulltext_keywords)

    if title_matched:
        return "sex-aware main focus", title_matched, full_matched
    if full_matched:
        return "sex-aware consideration", title_matched, full_matched
    return "", title_matched, full_matched


def update_workbook_sex_keywords(
    year_folder: Path,
    workbook_path: Path,
    log_rows: list[dict],
):
    """
    For each PDF in each month folder, detect sex keywords and
    update the workbook's sex-specific columns.
    """
    wb = openpyxl.load_workbook(workbook_path)

    for month in MONTHS:
        month_folder = year_folder / month
        if not month_folder.exists() or month not in wb.sheetnames:
            continue

        ws = wb[month]
        header = {cell.value: cell.column for cell in ws[1]}

        sex_kw_col = header.get("Sex-specific keywords?")
        sex_match_col = header.get("Sex keywords matched")
        sex_level_col = header.get("Sex-aware level")
        fn_col = header.get("Filename")
        title_col = header.get("Title")

        if not sex_kw_col or not sex_match_col:
            log.warning(f"  Sex keyword columns missing in sheet '{month}' — skipping.")
            continue

        if not sex_level_col:
            sex_level_col = ws.max_column + 1
            ws.cell(row=1, column=sex_level_col, value="Sex-aware level")

        # Build filename → row map
        fn_to_row: dict[str, int] = {}
        for row_idx in range(2, ws.max_row + 1):
            fn_val = ws.cell(row=row_idx, column=fn_col).value if fn_col else None
            if fn_val:
                fn_to_row[fn_val.lower()] = row_idx

        pdfs = sorted(month_folder.glob("*.pdf"))
        log.info(f"  Sex-scan: {len(pdfs)} PDFs in {month}…")

        for pdf in pdfs:
            text = extract_full_text_pypdf2(pdf)
            target_row = fn_to_row.get(pdf.name.lower())
            title_text = ""
            if target_row:
                # In this project workbook, article title may be stored in Filename.
                title_from_title_col = str(ws.cell(row=target_row, column=title_col).value or "") if title_col else ""
                title_from_filename = str(ws.cell(row=target_row, column=fn_col).value or "") if fn_col else ""
                title_text = title_from_title_col.strip() or title_from_filename.strip()

            level, title_matched, full_matched = classify_sex_aware_level(title_text, text)
            all_matched = sorted(set(title_matched) | set(full_matched))

            has_sex_analysis = "Yes" if all_matched else "No"
            matched_str = "; ".join(all_matched) if all_matched else ""

            # Find matching workbook row
            if target_row:
                ws.cell(row=target_row, column=sex_kw_col, value=has_sex_analysis)
                ws.cell(row=target_row, column=sex_match_col, value=matched_str)
                ws.cell(row=target_row, column=sex_level_col, value=level)
            else:
                log.warning(f"    No workbook row found for {pdf.name}")

            log_rows.append({
                "pdf": pdf.name,
                "month": month,
                "sex_analysis": has_sex_analysis,
                "sex_aware_level": level or "none",
                "title_keywords": "; ".join(title_matched) if title_matched else "none",
                "sex_keywords": matched_str,
                "n_keywords": len(all_matched),
            })

    wb.save(workbook_path)
    log.info(f"\nWorkbook updated with sex-keyword results.")


def main():
    parser = argparse.ArgumentParser(
        description="Scan A&D PDFs for sex-specific analysis keywords."
    )
    parser.add_argument("--year-folder", type=Path, required=True,
                        help="Folder containing month subfolders with PDFs")
    parser.add_argument("--xlsx", type=Path, required=True,
                        help="Path to the OSF-style Excel workbook")
    args = parser.parse_args()

    log.info("=== Step 4: Sex-specific keyword scan ===")
    log_rows: list[dict] = []

    update_workbook_sex_keywords(args.year_folder, args.xlsx, log_rows)

    # Save log
    log_path = args.xlsx.parent / "sex_keyword_scan_log.csv"
    with open(log_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "pdf",
                "month",
                "sex_analysis",
                "sex_aware_level",
                "title_keywords",
                "sex_keywords",
                "n_keywords",
            ],
        )
        writer.writeheader()
        writer.writerows(log_rows)

    total = len(log_rows)
    with_sex = sum(1 for r in log_rows if r["sex_analysis"] == "Yes")
    log.info(f"\nSex keyword scan summary:")
    log.info(f"  Total PDFs scanned : {total}")
    log.info(f"  Papers with sex keywords : {with_sex} ({100*with_sex/total:.1f}%)" if total else "  No PDFs processed.")
    log.info(f"  Log saved : {log_path}")
    log.info("=== Step 4 complete ===")


if __name__ == "__main__":
    main()
