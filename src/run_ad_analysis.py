"""
run_ad_analysis.py
-------------------
Step 6 of the AD Reproducibility Audit pipeline.

Reproduces key statistics from the MRM notebook, adapted for A&D:
- Total papers analysed
- Keyword match rate (open-science indicators)
- Code sharing rate
- Data sharing rate
- Hosting platform breakdown (GitHub, OSF, Zenodo, Dryad…)
- Country distribution of first authors
- Sex-specific keyword rate (NOVEL — not in MRM pipeline)
- Conditional sharing rates by country

Adapted from: KahinaBch/mrm-reproducible-research-2025
"""

import argparse
import logging
from collections import Counter
from pathlib import Path

import pandas as pd
import openpyxl
from scipy import stats

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

HOSTING_PATTERNS = {
    "GitHub": "github",
    "OSF": "osf.io",
    "Zenodo": "zenodo",
    "Dryad": "datadryad",
    "Figshare": "figshare",
    "Other": "",
}


def load_workbook_to_df(xlsx_path: Path) -> pd.DataFrame:
    """Load all month sheets from the workbook into a single DataFrame."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    frames = []
    for month in MONTHS:
        if month not in wb.sheetnames:
            continue
        ws = wb[month]
        data = list(ws.values)
        if len(data) < 2:
            continue
        header = data[0]
        rows = data[1:]
        df = pd.DataFrame(rows, columns=header)
        df["_Month"] = month
        # Excel sheet row numbers start at 1 and row 1 is the header.
        # Data rows therefore begin at 2.
        df["_RowInSheet"] = list(range(2, 2 + len(df)))
        frames.append(df)
    wb.close()
    if not frames:
        raise ValueError("No data found in workbook — run Steps 2–5 first.")
    return pd.concat(frames, ignore_index=True)


def _safe_yes_series(df: pd.DataFrame, col: str) -> pd.Series:
    """Return boolean Series for values that equal 'yes' (case-insensitive)."""
    if col not in df.columns:
        return pd.Series(False, index=df.index)
    s = df[col].fillna("").astype(str).str.lower().str.strip()
    return s.eq("yes")


def _looks_like_repo_link(val: str) -> bool:
    if not val:
        return False
    s = str(val).strip().lower()
    if not s or s in {"none", "nan"}:
        return False
    return s.startswith("http") or "github.com" in s or "gitlab.com" in s or "osf.io" in s or "zenodo" in s


def augment_from_keyword_scan_log(df: pd.DataFrame, keyword_log_csv: Path | None) -> pd.DataFrame:
    """Merge Step 3 keyword scan log into workbook-derived df.

    The log file is expected to have columns: month, matched_row, keywords_found, repo_link.
    Join key: (month, matched_row) == (df._Month, df._RowInSheet)

    We use it to fill missing 'Link' / 'Code repository link' and infer sharing flags.
    """
    if not keyword_log_csv:
        return df
    keyword_log_csv = Path(keyword_log_csv)
    if not keyword_log_csv.exists():
        return df

    log_df = pd.read_csv(keyword_log_csv)
    required = {"month", "matched_row", "keywords_found", "repo_link"}
    if not required.issubset(set(c.lower() for c in log_df.columns)):
        # Try case-sensitive fallback first
        cols = set(log_df.columns)
        if not required.issubset(cols):
            return df

    # Normalize column names
    log_df = log_df.rename(columns={
        "Month": "month",
        "matched_row": "matched_row",
        "Matched_Row": "matched_row",
        "keywords_found": "keywords_found",
        "Keywords_Found": "keywords_found",
        "repo_link": "repo_link",
        "Repo_Link": "repo_link",
    })

    log_df["month"] = log_df["month"].astype(str)
    log_df["matched_row"] = pd.to_numeric(log_df["matched_row"], errors="coerce")
    log_df = log_df.dropna(subset=["matched_row"])
    log_df["matched_row"] = log_df["matched_row"].astype(int)

    df2 = df.merge(
        log_df[["month", "matched_row", "keywords_found", "repo_link"]],
        how="left",
        left_on=["_Month", "_RowInSheet"],
        right_on=["month", "matched_row"],
    )

    # Fill link columns if present
    repo_link = df2.get("repo_link", pd.Series("", index=df2.index)).fillna("")
    if "Link" in df2.columns:
        link_existing = df2["Link"].fillna("").astype(str).str.strip()
        df2["Link"] = link_existing.where(link_existing.ne(""), repo_link)
    if "Code repository link" in df2.columns:
        repo_existing = df2["Code repository link"].fillna("").astype(str).str.strip()
        df2["Code repository link"] = repo_existing.where(repo_existing.ne(""), repo_link)

    # Infer sharing flags if missing/blank in workbook
    code_indicators = {
        "github", "gitlab", "repository", "repo", "script", "pipeline", "workflow", "open-source", "open source",
        "code availability", "reproducible", "reproducibility",
    }
    data_indicators = {"zenodo", "dryad", "figshare", "osf", "dataverse", "open data", "dataset", "data availability"}

    kw = df2.get("keywords_found", pd.Series("", index=df2.index)).fillna("").astype(str).str.lower()
    kw_set = kw.apply(lambda v: {t.strip() for t in v.split(";") if t.strip()})

    inferred_code = kw_set.apply(lambda s: any(t in s for t in code_indicators)) | repo_link.apply(_looks_like_repo_link)
    inferred_data = kw_set.apply(lambda s: any(t in s for t in data_indicators))

    if "Shared code?" in df2.columns:
        existing = df2["Shared code?"].fillna("").astype(str).str.strip()
        df2["Shared code?"] = existing.where(existing.ne(""), inferred_code.map(lambda b: "Yes" if b else ""))
    if "Shared data?" in df2.columns:
        existing = df2["Shared data?"].fillna("").astype(str).str.strip()
        df2["Shared data?"] = existing.where(existing.ne(""), inferred_data.map(lambda b: "Yes" if b else ""))

    return df2


def classify_hosting(link_val: str) -> str:
    """Classify a link string into a hosting platform."""
    if not link_val:
        return ""
    link_lower = str(link_val).lower()
    for platform, pattern in HOSTING_PATTERNS.items():
        if platform == "Other":
            continue
        if pattern in link_lower:
            return platform
    if link_lower.strip():
        return "Other"
    return ""


def run_analysis(df: pd.DataFrame, year: int) -> dict:
    """Compute all statistics. Returns dict of results."""
    results = {}

    # ── Basic counts ─────────────────────────────────────────────────────────
    total = len(df)
    results["total_papers"] = total
    results["year"] = year

    # Exclude rows where False Positive? == "Yes"
    df_valid = df[df.get("False Positive?", pd.Series(dtype=str)).str.lower() != "yes"].copy()
    results["total_valid"] = len(df_valid)

    # ── Open-science keyword match ───────────────────────────────────────────
    has_kw = df_valid["Keywords Matched"].notna() & (df_valid["Keywords Matched"] != "") & (df_valid["Keywords Matched"] != "none")
    results["n_keyword_match"] = int(has_kw.sum())
    results["pct_keyword_match"] = round(100 * has_kw.mean(), 1)

    # ── Code and data sharing ────────────────────────────────────────────────
    shared_code = _safe_yes_series(df_valid, "Shared code?")
    shared_data = _safe_yes_series(df_valid, "Shared data?")
    shared_any = shared_code | shared_data

    results["n_shared_code"] = int(shared_code.sum())
    results["n_shared_data"] = int(shared_data.sum())
    results["n_shared_any"] = int(shared_any.sum())
    results["pct_shared_code"] = round(100 * shared_code.mean(), 1)
    results["pct_shared_data"] = round(100 * shared_data.mean(), 1)
    results["pct_shared_any"] = round(100 * shared_any.mean(), 1)

    # ── Hosting platform ─────────────────────────────────────────────────────
    if "Link" in df_valid.columns:
        df_valid["_hosting"] = df_valid["Link"].apply(classify_hosting)
    else:
        df_valid["_hosting"] = ""
    hosting_counts = df_valid[df_valid["_hosting"] != ""]["_hosting"].value_counts().to_dict()
    results["hosting_counts"] = hosting_counts

    # ── Sex-specific keywords ─────────────────────────────────────────────────
    if "Sex-specific keywords?" in df_valid.columns:
        has_sex = df_valid["Sex-specific keywords?"].str.lower().str.strip() == "yes"
        results["n_sex_specific"] = int(has_sex.sum())
        results["pct_sex_specific"] = round(100 * has_sex.mean(), 1)

        # Most common sex keywords
        sex_kw_col = df_valid.get("Sex keywords matched", pd.Series(dtype=str))
        all_sex_kw = []
        for val in sex_kw_col.dropna():
            all_sex_kw.extend([k.strip() for k in str(val).split(";") if k.strip()])
        results["top_sex_keywords"] = Counter(all_sex_kw).most_common(10)

    # ── Country distribution ──────────────────────────────────────────────────
    if "First author affiliation country" in df_valid.columns:
        country_counts = df_valid["First author affiliation country"].dropna()
        country_counts = country_counts[country_counts != ""]
        results["country_counts"] = country_counts.value_counts().to_dict()
        results["n_country_identified"] = int((country_counts != "").sum())

        # Conditional sharing by country (countries with ≥5 papers)
        df_country = df_valid[df_valid["First author affiliation country"].notna()].copy()
        df_country = df_country[df_country["First author affiliation country"] != ""]
        df_country["_shared"] = (
            df_country["Shared code?"].str.lower().str.strip().eq("yes") |
            df_country["Shared data?"].str.lower().str.strip().eq("yes")
        )
        country_sharing = {}
        for country, grp in df_country.groupby("First author affiliation country"):
            if len(grp) >= 5:
                country_sharing[country] = {
                    "n": len(grp),
                    "n_shared": int(grp["_shared"].sum()),
                    "pct_shared": round(100 * grp["_shared"].mean(), 1),
                }
        results["country_sharing"] = country_sharing

        # Chi-square: country × sharing (for countries with ≥5 papers)
        if len(country_sharing) >= 2:
            eligible = [c for c, v in country_sharing.items() if v["n"] >= 5]
            df_chisq = df_country[df_country["First author affiliation country"].isin(eligible)]
            contingency = pd.crosstab(
                df_chisq["First author affiliation country"],
                df_chisq["_shared"]
            )
            if min(contingency.shape) > 1 and contingency.values.sum() > 0:
                chi2, p_val, dof, _ = stats.chi2_contingency(contingency)
                n_total = contingency.values.sum()
                denom = n_total * (min(contingency.shape) - 1)
                cramers_v = (chi2 / denom) ** 0.5 if denom else float("nan")
                results["country_chisq"] = {
                    "chi2": round(float(chi2), 3),
                    "p": round(float(p_val), 4),
                    "dof": int(dof),
                    "cramers_v": round(float(cramers_v), 3) if cramers_v == cramers_v else "N/A",
                }

    return results


def save_summary_excel(results: dict, out_path: Path, year: int):
    """Save statistical summary to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="6B21A8")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    rows = [
        ["Metric", "Value"],
        ["Year", results["year"]],
        ["Total papers in workbook", results["total_papers"]],
        ["Valid papers (excl. false positives)", results["total_valid"]],
        ["", ""],
        ["--- OPEN-SCIENCE KEYWORDS ---", ""],
        ["Papers with keyword match", results["n_keyword_match"]],
        ["% keyword match", f"{results['pct_keyword_match']}%"],
        ["", ""],
        ["--- CODE & DATA SHARING ---", ""],
        ["Shared code", results["n_shared_code"]],
        ["% shared code", f"{results['pct_shared_code']}%"],
        ["Shared data", results["n_shared_data"]],
        ["% shared data", f"{results['pct_shared_data']}%"],
        ["Shared code OR data", results["n_shared_any"]],
        ["% shared code OR data", f"{results['pct_shared_any']}%"],
        ["", ""],
        ["--- HOSTING PLATFORMS ---", ""],
    ]
    for platform, count in results.get("hosting_counts", {}).items():
        rows.append([f"  {platform}", count])

    rows += [
        ["", ""],
        ["--- SEX-SPECIFIC ANALYSIS ---", ""],
        ["Papers with sex-specific keywords", results.get("n_sex_specific", "N/A")],
        ["% papers with sex analysis", f"{results.get('pct_sex_specific', 'N/A')}%"],
        ["", ""],
        ["--- COUNTRY ---", ""],
        ["Papers with country identified", results.get("n_country_identified", "N/A")],
    ]

    if "country_chisq" in results:
        chisq = results["country_chisq"]
        rows += [
            ["Chi-square (country × sharing)", chisq["chi2"]],
            ["p-value", chisq["p"]],
            ["Cramér's V", chisq["cramers_v"]],
        ]

    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if i == 1:
                cell.font = header_font
                cell.fill = header_fill

    # Country sharing sheet
    if results.get("country_sharing"):
        ws2 = wb.create_sheet("Country Sharing")
        ws2.append(["Country", "N papers", "N shared", "% shared"])
        for country, vals in sorted(results["country_sharing"].items(),
                                     key=lambda x: x[1]["n"], reverse=True):
            ws2.append([country, vals["n"], vals["n_shared"], vals["pct_shared"]])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info(f"Summary saved: {out_path}")


def main():
    parser = argparse.ArgumentParser(description="Run statistical analysis on curated A&D workbook.")
    parser.add_argument("--xlsx", type=Path, required=True, help="Path to curated workbook")
    parser.add_argument("--year", type=int, required=True, help="Year analysed")
    parser.add_argument("--out", type=Path, default=None, help="Output Excel path")
    parser.add_argument(
        "--keyword-log-csv",
        type=Path,
        default=None,
        help="Optional Step 3 keyword scan log (default: alongside workbook if present)",
    )
    args = parser.parse_args()

    out_path = args.out or args.xlsx.with_name(f"AD_{args.year}_analysis.xlsx")

    log.info("=== Step 7: Statistical Analysis ===")
    df = load_workbook_to_df(args.xlsx)
    keyword_log = args.keyword_log_csv
    if keyword_log is None:
        candidate = args.xlsx.parent / "keyword_scan_log.csv"
        keyword_log = candidate if candidate.exists() else None
    df = augment_from_keyword_scan_log(df, keyword_log)
    log.info(f"  Loaded {len(df)} rows from workbook.")

    results = run_analysis(df, args.year)

    # Print key results
    log.info(f"\n{'='*50}")
    log.info(f"  Year: {results['year']} | Total papers: {results['total_papers']}")
    log.info(f"  Keyword match: {results['pct_keyword_match']}%")
    log.info(f"  Code sharing: {results['pct_shared_code']}%")
    log.info(f"  Data sharing: {results['pct_shared_data']}%")
    log.info(f"  Sex-specific: {results.get('pct_sex_specific', 'N/A')}%")
    log.info(f"{'='*50}\n")

    save_summary_excel(results, out_path, args.year)
    log.info("=== Step 7 complete ===")


if __name__ == "__main__":
    main()
