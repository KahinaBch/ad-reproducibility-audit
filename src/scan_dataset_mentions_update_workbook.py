"""
scan_dataset_mentions_update_workbook.py
---------------------------------------
Optional step: scan PDFs for dataset mentions using a controlled vocabulary.

This script reads dataset names from the AD dataset catalogue JSON
(`ad-dataset-catalogue/data/neuroimaging_genetics.json`, field: "name") and
searches each PDF's extracted text for those names (case-insensitive).

Outputs:
- Updates the workbook with two columns:
  - "Dataset(s) mentioned?" (Yes/No)
  - "Dataset names matched" (semicolon-separated dataset names)
- Writes a CSV log: workbooks/{year}/dataset_scan_log.csv

Notes:
- PDFs are not distributed in-repo (copyright).
- This is a keyword search heuristic; manual verification is recommended.
"""

import argparse
import csv
import json
import logging
import re
import unicodedata
from pathlib import Path

import openpyxl
import PyPDF2

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

MONTHS = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def default_dataset_json_path() -> Path:
    # Repo layout: files/ad-reproducibility-audit (this repo) next to files/ad-dataset-catalogue
    return _repo_root().parent / "ad-dataset-catalogue" / "data" / "neuroimaging_genetics.json"


def normalize_text(value: str) -> str:
    """Lowercase, strip accents, and replace non-alphanumerics with spaces."""
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def extract_text_pypdf2(pdf_path: Path) -> str:
    """Extract text from all pages using PyPDF2."""
    pages: list[str] = []
    try:
        with open(pdf_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                try:
                    pages.append(page.extract_text() or "")
                except Exception:
                    pages.append("")
    except Exception as exc:
        log.warning(f"  PyPDF2 failed on {pdf_path.name}: {exc}")
    return "\n".join(pages)


def load_dataset_names(dataset_json: Path) -> list[str]:
    """Load dataset names from the JSON catalogue (field: name)."""
    with open(dataset_json, "r", encoding="utf-8") as f:
        data = json.load(f)

    names: list[str] = []

    if isinstance(data, list):
        for item in data:
            if isinstance(item, dict):
                name = item.get("name")
                if isinstance(name, str) and name.strip():
                    names.append(name.strip())
    elif isinstance(data, dict):
        # Fallback: if the file structure changes, still attempt to find list-like fields.
        for value in data.values():
            if isinstance(value, list):
                for item in value:
                    if isinstance(item, dict):
                        name = item.get("name")
                        if isinstance(name, str) and name.strip():
                            names.append(name.strip())

    # Deduplicate, keep stable order
    seen = set()
    out = []
    for n in names:
        if n not in seen:
            seen.add(n)
            out.append(n)
    return out


def build_dataset_patterns(dataset_names: list[str]) -> list[tuple[str, re.Pattern]]:
    """Precompile search regex patterns against normalized text."""
    patterns: list[tuple[str, re.Pattern]] = []

    for raw_name in dataset_names:
        norm_name = normalize_text(raw_name)
        if not norm_name:
            continue

        # Skip extremely short names that are likely to create false positives.
        # Keep common acronyms like ADNI (length 4).
        if len(norm_name) <= 2:
            continue

        tokens = norm_name.split()
        if len(tokens) == 1:
            token = re.escape(tokens[0])
            pat = re.compile(rf"\b{token}\b")
        else:
            token_pat = r"\s+".join(rf"\b{re.escape(t)}\b" for t in tokens)
            pat = re.compile(token_pat)

        patterns.append((raw_name, pat))

    return patterns


def match_pdf_to_row(pdf_name: str, ws_rows: list[dict]) -> int | None:
    """Match a PDF filename to a workbook row. Returns 0-based index in ws_rows or None."""
    stem = Path(pdf_name).stem.lower().replace("-", "").replace("_", "").replace(" ", "")
    for i, row in enumerate(ws_rows):
        row_doi = str(row.get("DOI", "")).replace("/", "").replace(".", "").lower()
        if row_doi and row_doi in stem:
            return i
        row_fn = (
            str(row.get("Filename", ""))
            .lower()
            .replace("-", "")
            .replace("_", "")
            .replace(" ", "")
        )
        if row_fn and (stem in row_fn or row_fn in stem):
            return i
    return None


def scan_datasets(norm_text: str, patterns: list[tuple[str, re.Pattern]]) -> list[str]:
    matched: list[str] = []
    for raw_name, pat in patterns:
        if pat.search(norm_text):
            matched.append(raw_name)
    return matched


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Scan A&D PDFs for dataset mentions (based on dataset catalogue names) and update workbook."
    )
    parser.add_argument(
        "--year-folder",
        type=Path,
        required=True,
        help="Folder containing month subfolders with PDFs (e.g., data/raw/pdfs/2025)",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        required=True,
        help="Path to the Excel workbook (e.g., workbooks/2025/AD-ReproducibleResearch_2025.xlsx)",
    )
    parser.add_argument(
        "--dataset-json",
        type=Path,
        default=None,
        help="Path to neuroimaging_genetics.json (default: sibling repo ad-dataset-catalogue/data/neuroimaging_genetics.json)",
    )
    parser.add_argument(
        "--months",
        nargs="*",
        default=None,
        help="Optional subset of months to scan (e.g., January February). Default: all months.",
    )
    parser.add_argument(
        "--max-pdfs",
        type=int,
        default=None,
        help="Optional cap for number of PDFs per month (useful for quick tests).",
    )
    args = parser.parse_args()

    dataset_json = args.dataset_json or default_dataset_json_path()
    if not dataset_json.exists():
        raise FileNotFoundError(
            f"Dataset JSON not found: {dataset_json}. Provide --dataset-json explicitly."
        )

    dataset_names = load_dataset_names(dataset_json)
    patterns = build_dataset_patterns(dataset_names)

    log.info("=== Dataset mention scan: PDF keyword search ===")
    log.info(f"Dataset catalogue: {dataset_json}")
    log.info(f"Dataset names loaded: {len(dataset_names)}")

    wb = openpyxl.load_workbook(args.xlsx)
    months_to_scan = args.months or MONTHS

    log_rows: list[dict] = []

    for month in months_to_scan:
        month_folder = args.year_folder / month
        if not month_folder.exists():
            log.info(f"  Skipping {month} (folder not found)")
            continue
        if month not in wb.sheetnames:
            log.warning(f"  Skipping {month} (sheet not found in workbook)")
            continue

        ws = wb[month]
        header = {cell.value: cell.column for cell in ws[1]}

        mentioned_col = header.get("Dataset(s) mentioned?")
        if not mentioned_col:
            mentioned_col = ws.max_column + 1
            ws.cell(row=1, column=mentioned_col, value="Dataset(s) mentioned?")

        matched_col = header.get("Dataset names matched")
        if not matched_col:
            matched_col = ws.max_column + 1
            ws.cell(row=1, column=matched_col, value="Dataset names matched")

        # Build list of existing rows (for matching)
        existing_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {ws.cell(1, i + 1).value: v for i, v in enumerate(row)}
            existing_rows.append(row_dict)

        pdfs = sorted(month_folder.glob("*.pdf"))
        if args.max_pdfs is not None:
            pdfs = pdfs[: max(0, args.max_pdfs)]

        log.info(f"  Scanning {len(pdfs)} PDFs in {month}…")

        for pdf in pdfs:
            text = extract_text_pypdf2(pdf)
            norm_text = normalize_text(text)
            matched = scan_datasets(norm_text, patterns)

            row_idx = match_pdf_to_row(pdf.name, existing_rows)

            log_entry = {
                "pdf": pdf.name,
                "month": month,
                "datasets_found": "; ".join(matched) if matched else "none",
                "matched_row": row_idx + 2 if row_idx is not None else "unmatched",
            }
            log_rows.append(log_entry)

            if row_idx is None:
                log.warning(f"    Could not match PDF to workbook row: {pdf.name}")
                continue

            ws_row = row_idx + 2
            ws.cell(row=ws_row, column=mentioned_col, value="Yes" if matched else "No")
            ws.cell(row=ws_row, column=matched_col, value="; ".join(matched))

    wb.save(args.xlsx)

    log_path = args.xlsx.parent / "dataset_scan_log.csv"
    with open(log_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["pdf", "month", "datasets_found", "matched_row"])
        writer.writeheader()
        writer.writerows(log_rows)

    log.info(f"\nDataset scan log saved: {log_path}")
    log.info(f"Total PDFs scanned: {len(log_rows)}")
    log.info(f"PDFs with dataset mentions: {sum(1 for r in log_rows if r['datasets_found'] != 'none')}")
    log.info("=== Dataset scan complete ===")


if __name__ == "__main__":
    main()
