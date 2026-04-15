"""
scan_keywords_update_workbook.py
---------------------------------
Step 3 of the AD Reproducibility Audit pipeline.

Replicates the original MRM notebook's keyword-based screening logic:
- PyPDF2-based page-by-page text extraction
- Scans each PDF for code-availability and reproducibility keywords
- Updates the Excel workbook's "Keywords Matched" column
- Adds/updates a "Code repository link" column with detected repository URL

Adapted from: KahinaBch/mrm-reproducible-research-2025
"""

import argparse
import logging
import re
from pathlib import Path

import openpyxl
import PyPDF2

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Code availability + reproducibility keywords ─────────────────────────────
CODE_REPRO_KEYWORDS = [
    "code available",
    "code availability",
    "source code",
    "open source",
    "open-source",
    "open code",
    "code sharing",
    "supplementary code",
    "reproducible",
    "reproducibility",
    "reproducible research",
    "workflow",
    "pipeline",
    "script",
    "github",
    "gitlab",
    "bitbucket",
    "osf",
    "zenodo",
    "repository",
]

URL_PATTERN = re.compile(r"https?://[^\s)\]>'\"]+", re.IGNORECASE)
REPO_HOST_HINTS = ("github.com", "gitlab.com", "bitbucket.org", "osf.io", "zenodo.org")

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


def extract_text_pypdf2(pdf_path: Path) -> list[str]:
    """Extract text page by page using PyPDF2. Returns list of page texts."""
    pages = []
    try:
        with open(pdf_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                try:
                    pages.append(page.extract_text() or "")
                except Exception:
                    pages.append("")
    except Exception as e:
        log.warning(f"  PyPDF2 failed on {pdf_path.name}: {e}")
    return pages


def scan_keywords(pages: list[str]) -> list[str]:
    """
    Scan pages for code-availability and reproducibility keywords.
    Stops scanning pages once a keyword is found (replicates MRM notebook logic).
    Returns list of matched keywords.
    """
    matched = set()
    for keyword in CODE_REPRO_KEYWORDS:
        for page_text in pages:
            if keyword.lower() in page_text.lower():
                matched.add(keyword)
                break  # Stop scanning pages for this keyword once found
    return sorted(matched)


def extract_repository_link(pages: list[str]) -> str:
    """Extract first repository-like URL, prioritizing common code-hosting domains."""
    urls: list[str] = []
    for page_text in pages:
        if not page_text:
            continue
        urls.extend(URL_PATTERN.findall(page_text))

    if not urls:
        return ""

    cleaned = [u.rstrip(".,;)") for u in urls]
    for host in REPO_HOST_HINTS:
        for url in cleaned:
            if host in url.lower():
                return url
    return ""


def match_pdf_to_row(pdf_name: str, ws_rows: list[dict]) -> int | None:
    """
    Match a PDF filename to a workbook row.
    Strategy: DOI match first, then filename heuristic.
    Returns 1-based row index or None.
    """
    stem = Path(pdf_name).stem.lower().replace("-", "").replace("_", "").replace(" ", "")
    for i, row in enumerate(ws_rows):
        # DOI-based match
        row_doi = str(row.get("DOI", "")).replace("/", "").replace(".", "").lower()
        if row_doi and row_doi in stem:
            return i
        # Filename heuristic
        row_fn = str(row.get("Filename", "")).lower().replace("-", "").replace("_", "").replace(" ", "")
        if row_fn and (stem in row_fn or row_fn in stem):
            return i
    return None


def process_month_folder(
    month_folder: Path,
    workbook_path: Path,
    month_name: str,
    log_rows: list[dict],
):
    """Scan all PDFs in a month folder and update the corresponding workbook sheet."""
    wb = openpyxl.load_workbook(workbook_path)

    if month_name not in wb.sheetnames:
        log.warning(f"  Sheet '{month_name}' not found in workbook.")
        return

    ws = wb[month_name]

    # Build column index map
    header = {cell.value: cell.column for cell in ws[1]}
    kw_col = header.get("Keywords Matched")
    if not kw_col:
        log.error("  Column 'Keywords Matched' not found in sheet.")
        return

    repo_col = header.get("Code repository link")
    if not repo_col:
        repo_col = ws.max_column + 1
        ws.cell(row=1, column=repo_col, value="Code repository link")

    # Build list of existing rows (for matching)
    existing_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {ws.cell(1, i + 1).value: v for i, v in enumerate(row)}
        existing_rows.append(row_dict)

    pdfs = sorted(month_folder.glob("*.pdf"))
    log.info(f"  Processing {len(pdfs)} PDFs in {month_folder.name}…")

    for pdf in pdfs:
        pages = extract_text_pypdf2(pdf)
        keywords = scan_keywords(pages)
        repo_link = extract_repository_link(pages)

        row_idx = match_pdf_to_row(pdf.name, existing_rows)

        log_entry = {
            "pdf": pdf.name,
            "month": month_name,
            "keywords_found": "; ".join(keywords) if keywords else "none",
            "repo_link": repo_link or "none",
            "matched_row": row_idx + 2 if row_idx is not None else "unmatched",
        }
        log_rows.append(log_entry)

        if row_idx is not None:
            ws.cell(row=row_idx + 2, column=kw_col, value="; ".join(keywords))
            ws.cell(row=row_idx + 2, column=repo_col, value=repo_link)
        else:
            log.warning(f"    Could not match PDF to workbook row: {pdf.name}")

    wb.save(workbook_path)
    log.info(f"  Workbook saved after {month_name}.")


def main():
    parser = argparse.ArgumentParser(
        description="Scan A&D PDFs for code-availability/reproducibility keywords and update workbook."
    )
    parser.add_argument("--year-folder", type=Path, required=True,
                        help="Folder containing month subfolders with PDFs")
    parser.add_argument("--xlsx", type=Path, required=True,
                        help="Path to the OSF-style Excel workbook")
    args = parser.parse_args()

    log.info("=== Step 3: Code availability keyword scan ===")
    log_rows = []

    for month in MONTHS:
        month_folder = args.year_folder / month
        if month_folder.exists():
            process_month_folder(month_folder, args.xlsx, month, log_rows)
        else:
            log.info(f"  Skipping {month} (folder not found)")

    # Save keyword scan log
    import csv
    log_path = args.xlsx.parent / "keyword_scan_log.csv"
    with open(log_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["pdf", "month", "keywords_found", "repo_link", "matched_row"])
        writer.writeheader()
        writer.writerows(log_rows)

    log.info(f"\nKeyword scan log saved: {log_path}")
    log.info(f"Total PDFs scanned: {len(log_rows)}")
    log.info(f"PDFs with keywords: {sum(1 for r in log_rows if r['keywords_found'] != 'none')}")
    log.info("=== Step 3 complete ===")


if __name__ == "__main__":
    main()
